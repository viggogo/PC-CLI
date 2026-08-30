import os
from datetime import date

import openpyxl
import pytest

from hevy_mcp import excel_writer as ew


@pytest.fixture
def tmp_book(tmp_path, monkeypatch):
    path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Træning"
    headers = ["Date", "Place", "Type", "Time", "Mave", "AddCardio",
               "AddCardio2", "Total", "Ensamble", "Rating", "Claude", "Comments"]
    ws.append(headers)
    ws.append([date(2026, 5, 1), "AC", "push", 50, 0, 0, 0,
               "=SUM(D2:G2)", 0, 3, 1, "seed"])
    wb.save(path)
    monkeypatch.setenv("EXCEL_PATH", str(path))
    monkeypatch.setenv("EXCEL_SHEET", "Træning")
    return path


def test_total_formula_text():
    assert ew.total_formula(5) == (
        '=IF(ISBLANK(Table4[[#This Row],[Date]]), " ", '
        'SUM(D5:G5)-IF(E5 = 1, 1, 0))')


def test_append_writes_values_claude_and_total_formula(tmp_book):
    row = {
        "Date": date(2026, 7, 4), "Place": "AC", "Type": "upper", "Time": 79,
        "Mave": 1, "AddCardio": 15, "AddCardio2": 0, "Ensamble": 0,
        "Rating": 3, "Claude": 1, "Comments": "3",
    }
    n = ew.append_rows([row])
    assert n == 1
    wb = openpyxl.load_workbook(tmp_book)
    ws = wb["Træning"]
    r = 3  # header + seed + this
    assert ws.cell(row=r, column=ew.COLUMN_INDEX["Type"]).value == "upper"
    assert ws.cell(row=r, column=ew.COLUMN_INDEX["Time"]).value == 79
    assert ws.cell(row=r, column=ew.COLUMN_INDEX["AddCardio"]).value == 15
    assert ws.cell(row=r, column=ew.COLUMN_INDEX["Rating"]).value == 3
    assert ws.cell(row=r, column=ew.COLUMN_INDEX["Claude"]).value == 1
    assert ws.cell(row=r, column=ew.COLUMN_INDEX["Comments"]).value == "3"
    # Total is the formula, not a number:
    assert ws.cell(row=r, column=ew.COLUMN_INDEX["Total"]).value == ew.total_formula(r)


def test_append_blank_rating_is_empty(tmp_book):
    row = {
        "Date": date(2026, 7, 5), "Place": "", "Type": "pull", "Time": 40,
        "Mave": 0, "AddCardio": 0, "AddCardio2": 0, "Ensamble": 0,
        "Rating": None, "Claude": 1, "Comments": "",
    }
    ew.append_rows([row])
    wb = openpyxl.load_workbook(tmp_book)
    ws = wb["Træning"]
    r = 3
    assert ws.cell(row=r, column=ew.COLUMN_INDEX["Rating"]).value is None
    assert ws.cell(row=r, column=ew.COLUMN_INDEX["Place"]).value is None



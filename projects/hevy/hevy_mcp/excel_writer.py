import os
from datetime import date, datetime
from typing import Optional

import openpyxl

# Physical layout AFTER the Claude-column migration (1-based indices).
COLUMN_INDEX = {
    "Date": 1, "Place": 2, "Type": 3, "Time": 4, "Mave": 5,
    "AddCardio": 6, "AddCardio2": 7, "Total": 8, "Ensamble": 9,
    "Rating": 10, "Claude": 11, "Comments": 12,
}

# Value columns the mapper supplies. Total is a formula, not a value.
_VALUE_COLUMNS = [
    "Date", "Place", "Type", "Time", "Mave", "AddCardio", "AddCardio2",
    "Ensamble", "Rating", "Claude", "Comments",
]


def total_formula(r: int) -> str:
    return (f'=IF(ISBLANK(Table4[[#This Row],[Date]]), " ", '
            f'SUM(D{r}:G{r})-IF(E{r} = 1, 1, 0))')


def _open() -> tuple[openpyxl.Workbook, object]:
    path = os.environ["EXCEL_PATH"]
    sheet = os.environ.get("EXCEL_SHEET", "Træning")
    wb = openpyxl.load_workbook(path)
    return wb, wb[sheet]


def get_last_date() -> Optional[date]:
    """Return the date in the last non-empty row of the sheet."""
    _, ws = _open()
    last_date: Optional[date] = None
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val is None:
            continue
        if isinstance(val, datetime):
            last_date = val.date()
        elif isinstance(val, date):
            last_date = val
    return last_date


def _find_first_empty_row(ws) -> int:
    """First row index (1-based) with no Date value."""
    for i in range(2, ws.max_row + 2):
        if ws.cell(row=i, column=1).value is None:
            return i
    return ws.max_row + 1


def append_rows(rows: list[dict]) -> int:
    """Append row dicts to the Træning sheet. Returns count written."""
    path = os.environ["EXCEL_PATH"]
    sheet = os.environ.get("EXCEL_SHEET", "Træning")
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]

    first_empty = _find_first_empty_row(ws)

    for idx, row_data in enumerate(rows):
        r = first_empty + idx
        for col_name in _VALUE_COLUMNS:
            value = row_data.get(col_name)
            ws.cell(row=r, column=COLUMN_INDEX[col_name],
                    value=value if value not in ("", None) else None)
        ws.cell(row=r, column=COLUMN_INDEX["Total"], value=total_formula(r))

    wb.save(path)
    return len(rows)



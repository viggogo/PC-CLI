"""All workbook I/O for the Sengetøj sheet.

Two facts about the real sheet drive everything here:

  * Table2's ref is A1:B1048576 and ws.max_row is 1048576, because the Diff
    formula is materialized in every one of those rows. Neither says where
    the data ends. Only a scan of column A does.

  * That also makes a plain load_workbook() cost ~6.6 seconds. Reading goes
    through read_only mode over column A and breaks as soon as it sees a run
    of blanks, which costs ~0.01 seconds. A read_only walk that does NOT
    break early costs ~4 seconds, so the break is not an optimisation — it
    is the difference between a usable tool and an unusable one.
"""

import datetime as dt
from pathlib import Path
from typing import Iterable

import openpyxl

# How many consecutive empty cells in column A mean "the data ended".
# Comfortably more than any gap a hand-kept column could contain.
BLANK_RUN = 100

DATE_COL = 1
CLI_COL = 3
CLI_HEADER = "cli"

# Row 1 is the header, so data starts here.
FIRST_DATA_ROW = 2


class SheetMissing(Exception):
    """The configured tab is not in the workbook."""


def _as_date(value) -> dt.date:
    return value.date() if isinstance(value, dt.datetime) else value


def scan(values: Iterable) -> tuple[list[dt.date], int]:
    """Turn column A values (starting at row 2) into (dates, first_empty_row).

    Stops after BLANK_RUN consecutive blanks, so a caller may hand this a
    lazy iterator over a million rows without paying for them.
    """
    dates: list[dt.date] = []
    last_row = FIRST_DATA_ROW - 1
    blanks = 0
    row = FIRST_DATA_ROW

    for value in values:
        if value is None:
            blanks += 1
            if blanks >= BLANK_RUN:
                break
        else:
            blanks = 0
            dates.append(_as_date(value))
            last_row = row
        row += 1

    return dates, last_row + 1


def read_dates(path: Path, sheet_name: str) -> list[dt.date]:
    """Every recorded change, ascending. Fast, and safe while Excel is open."""
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise SheetMissing(sheet_name)
        ws = wb[sheet_name]
        # ws.cell() is unsupported in read_only mode; iter_rows is the only way.
        values = (row[0] for row in
                  ws.iter_rows(min_row=FIRST_DATA_ROW, max_col=DATE_COL,
                               values_only=True))
        dates, _ = scan(values)
        return dates
    finally:
        wb.close()


def gaps(dates: list[dt.date]) -> list[int | None]:
    """Days since the previous entry, aligned to `dates`. None for the first."""
    return [None if i == 0 else (d - dates[i - 1]).days
            for i, d in enumerate(dates)]


class WorkbookLocked(Exception):
    """Excel is holding the file open, so it cannot be saved."""


def open_for_write(path: Path, sheet_name: str):
    """Load the workbook for editing. Costs ~6.6s on the real file.

    read_only mode cannot be used here: it forbids ws.cell() and cannot save.
    """
    wb = openpyxl.load_workbook(path)
    if sheet_name not in wb.sheetnames:
        # close() is a no-op outside read_only/write_only mode, so this cannot
        # leak a handle and no test can pin it. Kept because it is correct
        # and would matter if this ever moved off normal mode. The path
        # where it IS load-bearing is read_dates(), which uses
        # read_only=True -- see its release tests in test_sheet_read.py.
        wb.close()
        raise SheetMissing(sheet_name)
    return wb, wb[sheet_name]


def dates_and_next_row(ws) -> tuple[list[dt.date], int]:
    """The same scan as read_dates, over an already-open worksheet.

    iter_rows rather than ws.cell(): it reads the column cleanly and, via
    scan()'s blank-run break, stops early instead of walking to ws.max_row.
    In normal mode it does NOT avoid creating cells -- openpyxl 3.1.5's
    Worksheet._cells_by_row is a generator over ws.cell(row=..., column=...),
    so every cell iter_rows visits (including blanks) is instantiated, same
    as calling ws.cell() directly would. That is harmless here: openpyxl's
    writer skips any cell with no value and no style, so those phantom
    blanks are never persisted and the saved file does not grow. iter_rows
    is also bounded by ws.max_row, which correctly yields nothing at all on
    an empty sheet.
    """
    values = (row[0] for row in
              ws.iter_rows(min_row=FIRST_DATA_ROW, max_col=DATE_COL,
                           values_only=True))
    return scan(values)


def has_cli_column(ws) -> bool:
    # ws._cells.get(...) rather than ws.cell(1, CLI_COL): the latter CREATES
    # the cell if it is absent, which on the real, pre-migration file is a
    # phantom C1 -- exactly the hazard migrate._peek exists to avoid.
    cell = ws._cells.get((1, CLI_COL))
    header = cell.value if cell is not None else None
    return isinstance(header, str) and header.strip().lower() == CLI_HEADER


def append_entry(ws, when: dt.date, row: int) -> None:
    """Write the date and the cli flag. Two cells, nothing else.

    Column B already holds the Diff formula in this row and every row below,
    and Table2's ref already spans the full column, so neither needs touching.
    """
    target = ws.cell(row, DATE_COL)
    target.value = dt.datetime(when.year, when.month, when.day)
    # Only copy from a preceding DATA row. At FIRST_DATA_ROW the row above is
    # the header, whose format is meaningless here -- copying its General
    # format onto a date cell renders the date as a serial number.
    if row - 1 >= FIRST_DATA_ROW:
        above = ws.cell(row - 1, DATE_COL)
        if above.value is not None:
            target.number_format = above.number_format
    ws.cell(row, CLI_COL, 1)


def save(wb, path: Path) -> None:
    try:
        wb.save(path)
    except PermissionError as exc:
        raise WorkbookLocked(path) from exc
    finally:
        # close() is a no-op outside read_only/write_only mode, so this
        # cannot leak a handle and no test can pin it. Kept because it is
        # correct and would matter if this ever moved off normal mode. The
        # path where it IS load-bearing is read_dates(), which uses
        # read_only=True -- see its release tests in test_sheet_read.py.
        wb.close()

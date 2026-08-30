"""One-time restructuring of the Sengetøj sheet.

Two jobs, one save:

  1. Move the free-text comments one column right, D -> E. Dropping the new
     `cli` column into C would otherwise consume the blank column that
     separates the table from the comments, leaving the table butting
     straight up against them.

  2. Add the `cli` column to Table2 -- header, ref, autofilter and
     tableColumns, all of which must agree or Excel calls the file corrupt.

Run once, deliberately, never from the CLI. It takes a backup first.
"""

import argparse
import datetime as dt
import shutil
import sys
from copy import copy
from pathlib import Path

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import TableColumn

from . import sheet
from .env import excel_path, load_env, sheet_name

COMMENT_COL_BEFORE = 4  # D
COMMENT_COL_AFTER = 5   # E

TABLE_NAME = "Table2"


def comment_rows(ws, col: int) -> list[int]:
    """Every populated row in `col`, exactly.

    Enumerates ws._cells rather than walking a row window: the comments are
    sparse and sit both above and beside the data, so any window is a guess,
    and a comment outside it would be silently left behind by a migration
    that only runs once. This also avoids ws.cell(), which CREATES empty
    cells as a side effect of reading them.
    """
    return sorted(row for (row, column), cell in ws._cells.items()
                  if column == col and cell.value is not None)


def move_comments(ws) -> int:
    """Move every populated comment one column right. Returns the count.

    Deliberately cell-by-cell rather than ws.move_range: move_range with
    translate=True would rewrite the counter's INDEX(A:A,COUNTA(A:A)) into
    INDEX(B:B,COUNTA(B:B)), and even with translate=False it would not
    rebuild the ArrayFormula's own ref.

    Two guards, because this runs once against an irreplaceable file and
    must fail loudly rather than silently mangle it:
      - a populated destination cell stops the move before it overwrites
        anything;
      - a non-empty source column after the loop means comment_rows missed
        something, and that is raised rather than swallowed.
    """
    rows = comment_rows(ws, COMMENT_COL_BEFORE)

    for row in rows:
        src = ws.cell(row, COMMENT_COL_BEFORE)
        dst = ws.cell(row, COMMENT_COL_AFTER)

        if dst.value is not None:
            raise RuntimeError(
                f"kolonne E har allerede en værdi i række {row}: "
                f"{dst.value!r} -- flytning afbrudt for ikke at overskrive")

        value = src.value
        if isinstance(value, ArrayFormula):
            # The ref names the cell the formula lives in, so it moves too.
            value = ArrayFormula(ref=f"{dst.column_letter}{row}", text=value.text)

        dst.value = value
        dst._style = copy(src._style)

        src.value = None
        # Reset rather than leave the old style behind: the counter cell is
        # bold on a yellow fill, and an emptied cell that kept that fill
        # would show as a stray coloured block in the spacer column.
        # "Normal" is openpyxl's built-in default and clears both.
        src.style = "Normal"

    leftover = comment_rows(ws, COMMENT_COL_BEFORE)
    if leftover:
        raise RuntimeError(
            f"kolonne D er ikke tom efter flytningen: rækker {leftover}")

    before_letter = ws.cell(1, COMMENT_COL_BEFORE).column_letter
    after_letter = ws.cell(1, COMMENT_COL_AFTER).column_letter
    if before_letter in ws.column_dimensions:
        src_dim = ws.column_dimensions[before_letter]
        dst_dim = ws.column_dimensions[after_letter]
        # Width, visibility and outline grouping travel with the comments.
        # Not `style`: that is an index into the workbook's shared style
        # table, and copying it across columns is riskier than it is worth.
        dst_dim.width = src_dim.width
        dst_dim.hidden = src_dim.hidden
        dst_dim.outlineLevel = src_dim.outlineLevel
        dst_dim.bestFit = src_dim.bestFit
        del ws.column_dimensions[before_letter]

    return len(rows)


def add_cli_column(ws) -> None:
    """Add `cli` to Table2. All four pieces must agree or Excel objects."""
    ws.cell(1, sheet.CLI_COL, sheet.CLI_HEADER)

    table = ws.tables[TABLE_NAME]
    # range_boundaries handles both edges -- not a hardcoded "A1" origin,
    # and not fragile against absolute ($) anchors the way stripping
    # non-digit characters out of the ref string would be.
    min_col, min_row, _, max_row = range_boundaries(table.ref)
    origin = f"{get_column_letter(min_col)}{min_row}"
    new_ref = f"{origin}:{get_column_letter(sheet.CLI_COL)}{max_row}"

    table.ref = new_ref
    if table.autoFilter is not None:
        table.autoFilter.ref = new_ref

    next_id = max((c.id for c in table.tableColumns), default=0) + 1
    table.tableColumns.append(TableColumn(id=next_id, name=sheet.CLI_HEADER))


def backup(path: Path) -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    target = path.with_name(f"{path.stem}.{stamp}.bak{path.suffix}")
    shutil.copy2(path, target)
    return target


def plan(ws) -> list[str]:
    """What still needs doing. Empty means the sheet is already migrated."""
    steps = []
    if comment_rows(ws, COMMENT_COL_BEFORE):
        steps.append(f"Flyt kommentarer D -> E "
                     f"({len(comment_rows(ws, COMMENT_COL_BEFORE))} celler)")
    if not sheet.has_cli_column(ws):
        steps.append(f'Tilføj kolonnen "{sheet.CLI_HEADER}" til {TABLE_NAME} (C)')
    return steps


def _peek(ws, row: int, col: int):
    """Read a cell's value without creating it as a side effect.

    ws.cell(row, col) CREATES the cell if it is absent -- the exact hazard
    comment_rows' own docstring names. On the real file C1 does not exist
    before migration, so asking "what does C1 hold" must not itself bring
    C1 into existence.
    """
    cell = ws._cells.get((row, col))
    return cell.value if cell is not None else None


def _blocked(ws) -> str | None:
    """A reason to refuse, or None."""
    header = _peek(ws, 1, sheet.CLI_COL)
    if header is not None and not sheet.has_cli_column(ws):
        return (f"C1 indeholder allerede {header!r}, ikke "
                f"{sheet.CLI_HEADER!r}. Migreringen er afbrudt.")
    if comment_rows(ws, COMMENT_COL_BEFORE) and comment_rows(ws, COMMENT_COL_AFTER):
        return ("Kolonne E er ikke tom, og kommentarerne står stadig i D. "
                "Migreringen er afbrudt.")
    return None


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m sengetoej.migrate",
        description="Engangsmigrering: flyt kommentarer D -> E og "
                    'tilføj kolonnen "cli" til Table2.')
    parser.add_argument("-y", "--yes", action="store_true",
                        help="spring bekræftelsen over")
    args = parser.parse_args(argv)

    load_env()
    path, tab = excel_path(), sheet_name()

    try:
        wb, ws = sheet.open_for_write(path, tab)
    except FileNotFoundError:
        print(f"Fejl: filen findes ikke: {path}", file=sys.stderr)
        return 1
    except sheet.SheetMissing:
        print(f"Fejl: arket {tab!r} findes ikke i {path}", file=sys.stderr)
        return 1

    reason = _blocked(ws)
    if reason:
        print(f"Fejl: {reason}", file=sys.stderr)
        wb.close()
        return 1

    steps = plan(ws)
    if not steps:
        print("Arket er allerede migreret. Intet at gøre.")
        wb.close()
        return 0

    print(f"{path}  [{tab}]")
    for step in steps:
        print(f"  - {step}")

    if not args.yes and input("Udfør migreringen? [y/N] ").strip().lower() not in ("y", "yes"):
        print("Afbrudt.")
        wb.close()
        return 0

    made = backup(path)
    print(f"Sikkerhedskopi: {made.name}")

    # Only run each job when it is actually outstanding -- move_comments'
    # trailing column-dimension block runs even when there is nothing to
    # move, and would otherwise silently overwrite E's width with D's (or
    # delete a width the user set on D since the last migration) on a run
    # whose printed plan never mentioned touching either column.
    if comment_rows(ws, COMMENT_COL_BEFORE):
        try:
            moved = move_comments(ws)
        except RuntimeError as exc:
            # Half-moved in memory only -- never save from here. The user
            # has just been told about the backup; point them at it.
            print(f"Fejl under flytning af kommentarer: {exc}", file=sys.stderr)
            print(f"Arket er IKKE gemt. Sikkerhedskopien ligger her: {made}",
                  file=sys.stderr)
            wb.close()
            return 1
    else:
        moved = 0

    added_cli = not sheet.has_cli_column(ws)
    if added_cli:
        add_cli_column(ws)

    try:
        sheet.save(wb, path)
    except sheet.WorkbookLocked:
        print("Fejl: Luk Excel og prøv igen.", file=sys.stderr)
        wb.close()
        return 1

    # Report only the work actually done -- on a one-shot migration this
    # line is the user's only record of what changed.
    done = []
    if moved:
        done.append(f"{moved} kommentar(er) flyttet")
    if added_cli:
        done.append(f'kolonnen "{sheet.CLI_HEADER}" tilføjet')
    print("Færdig. " + ", ".join(done) + ".")
    return 0


if __name__ == "__main__":
    sys.exit(main())

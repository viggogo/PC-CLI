import openpyxl
import pytest
from openpyxl.worksheet.formula import ArrayFormula

from sengetoej import migrate, sheet

from .conftest import COUNTER_TEXT, DATES, SHEET


def test_comment_rows_finds_every_populated_row(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    assert migrate.comment_rows(ws, migrate.COMMENT_COL_BEFORE) == [1, 2, 3, 4]


def test_comment_rows_is_empty_when_the_column_is(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    assert migrate.comment_rows(ws, migrate.COMMENT_COL_AFTER) == []


def test_comment_rows_finds_a_row_far_below_the_last_data_row(pre_migration_book):
    """The old implementation walked range(1, next_row + BLANK_RUN). With
    next_row ~5 and BLANK_RUN=100 that window ends around row 105, so a
    comment at row 200 sat outside it and was silently missed -- this test
    fails against that windowed version and passes against the exact
    ws._cells enumeration."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    ws.cell(200, migrate.COMMENT_COL_BEFORE, "Stray note")
    assert migrate.comment_rows(ws, migrate.COMMENT_COL_BEFORE) == [1, 2, 3, 4, 200]


def test_move_comments_finds_and_moves_a_comment_below_the_last_data_row(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    ws.cell(200, migrate.COMMENT_COL_BEFORE, "Stray note")
    moved = migrate.move_comments(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert moved == 5
    assert ws2.cell(200, migrate.COMMENT_COL_AFTER).value == "Stray note"
    assert ws2.cell(200, migrate.COMMENT_COL_BEFORE).value is None


def test_move_comments_relocates_values_to_column_e(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    moved = migrate.move_comments(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert moved == 4
    assert ws2["E1"].value == "Dage siden sidste skift"
    assert ws2["E3"].value == "Obs: Ferie"
    assert ws2["E4"].value == "Dato indikerer at der er blevet skiftet sengetøj"


def test_move_comments_empties_column_d(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.move_comments(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert [ws2.cell(r, 4).value for r in range(1, 10)] == [None] * 9


def test_move_comments_raises_rather_than_overwriting_column_e(pre_migration_book):
    """Nothing in the real sheet populates E before the move, but the guard
    exists so an unexpected value there stops the migration instead of
    silently overwriting it."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    ws.cell(3, migrate.COMMENT_COL_AFTER, "Already something here")

    with pytest.raises(RuntimeError):
        migrate.move_comments(ws)

    # And it must not have overwritten the pre-existing value on the way.
    assert ws.cell(3, migrate.COMMENT_COL_AFTER).value == "Already something here"


def test_move_comments_raises_if_column_d_is_not_empty_afterwards(pre_migration_book, monkeypatch):
    """Defensive post-condition: if comment_rows ever misses a populated
    cell, move_comments must fail loudly rather than leave it behind
    silently. Simulated by making the *first* call to comment_rows (the one
    move_comments uses to decide what to move) report one row short of the
    truth, so a real value is left in D when the loop finishes; the second
    call (the post-condition check) sees the real, complete state and must
    raise."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    real_comment_rows = migrate.comment_rows
    calls = {"n": 0}

    def flaky(ws_, col):
        calls["n"] += 1
        rows = real_comment_rows(ws_, col)
        if calls["n"] == 1 and col == migrate.COMMENT_COL_BEFORE:
            return rows[:-1]
        return rows

    monkeypatch.setattr(migrate, "comment_rows", flaky)

    with pytest.raises(RuntimeError):
        migrate.move_comments(ws)


def test_move_comments_does_not_translate_the_counter_formula(pre_migration_book):
    # The whole point: a translated move would make this INDEX(B:B,COUNTA(B:B)).
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.move_comments(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    counter = ws2["E2"].value
    assert isinstance(counter, ArrayFormula)
    assert counter.text == COUNTER_TEXT
    assert "A:A" in counter.text
    assert "B:B" not in counter.text


def test_move_comments_rebuilds_the_array_formula_ref(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.move_comments(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert ws2["E2"].value.ref == "E2"


def test_move_comments_carries_the_styling(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.move_comments(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert ws2["E2"].font.b is True
    # openpyxl normalises an RGB fill to 8 hex digits on round-trip, so
    # compare the colour, not the string.
    assert ws2["E2"].fill.fgColor.rgb.endswith("FFFF00")


def test_move_comments_moves_the_column_width(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.move_comments(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert ws2.column_dimensions["E"].width == pytest.approx(20.6)


def test_move_comments_resets_column_d_width_to_default(pre_migration_book):
    """The spacer column must not stay as wide as the comment column was --
    del ws.column_dimensions["D"] must actually run, not be a no-op. Compare
    against column F, which was never touched by anything, rather than a
    hardcoded number: whatever openpyxl considers "no explicit width" is
    what both should read as."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.move_comments(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert ws2.column_dimensions["D"].width == ws2.column_dimensions["F"].width
    assert ws2.column_dimensions["D"].width != pytest.approx(20.6)


def test_move_comments_carries_hidden_outline_and_bestfit(pre_migration_book):
    """Width is not the only column property that should travel: hidden,
    outlineLevel and bestFit move too, but NOT the style index."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    d_dim = ws.column_dimensions["D"]
    d_dim.hidden = True
    d_dim.outlineLevel = 2
    d_dim.bestFit = True

    migrate.move_comments(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    e_dim = ws2.column_dimensions["E"]
    assert e_dim.hidden is True
    assert e_dim.outlineLevel == 2
    assert e_dim.bestFit is True


def test_move_comments_leaves_the_dates_and_formulas_alone(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.move_comments(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert sheet.read_dates(pre_migration_book, SHEET) == [d.date() for d in DATES]
    for row in range(2, 5):
        assert ws2.cell(row, sheet.DATE_COL).number_format == "mm-dd-yy"
    assert "Table2[[#This Row],[Date]]" in ws2["B2"].value
    assert "Table2[[#This Row],[Date]]" in ws2["B60"].value


def test_move_comments_clears_column_d_styling(pre_migration_book):
    """D2 was bold on a yellow fill; after the move it must not be a stray
    coloured block. This pins the required deviation from the brief: reset
    via ws.style = "Normal" rather than borrowing column F's style (which
    would create a phantom cell in F)."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.move_comments(ws)

    # Column F must never have been touched -- ws.cell() on an untouched
    # cell creates it, so checking this on the live worksheet (rather than
    # after a save/reload round-trip, which may silently drop untouched
    # default cells) is the reliable way to catch the phantom-cell hazard.
    assert (1, 6) not in ws._cells and (2, 6) not in ws._cells

    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    d2 = ws2["D2"]
    assert d2.font.b is not True
    assert d2.fill.fgColor.rgb != "00FFFF00"

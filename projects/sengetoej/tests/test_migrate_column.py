import openpyxl
import pytest

from sengetoej import migrate, sheet

from .conftest import SHEET, TABLE_BOTTOM


def _spy_on_save(monkeypatch):
    """Track every call to sheet.save without changing its behaviour.

    A read_bytes() comparison alone is NOT reliable here: openpyxl's output
    only differs from an unmodified re-save once docProps/core.xml's
    dcterms:modified timestamp (second resolution) ticks over, so a fast
    test run can re-save the file and still see identical bytes. Call
    tracking is deterministic regardless of timing.
    """
    calls = []
    real_save = sheet.save

    def spy(wb, path):
        calls.append(path)
        return real_save(wb, path)

    monkeypatch.setattr(sheet, "save", spy)
    return calls


def test_add_cli_column_writes_the_header(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.add_cli_column(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert ws2["C1"].value == "cli"


def test_add_cli_column_extends_the_ref_and_the_autofilter(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.add_cli_column(ws)
    sheet.save(wb, pre_migration_book)

    table = openpyxl.load_workbook(pre_migration_book)[SHEET].tables["Table2"]
    assert table.ref == f"A1:C{TABLE_BOTTOM}"
    assert table.autoFilter.ref == f"A1:C{TABLE_BOTTOM}"


def test_add_cli_column_registers_the_table_column(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.add_cli_column(ws)
    sheet.save(wb, pre_migration_book)

    table = openpyxl.load_workbook(pre_migration_book)[SHEET].tables["Table2"]
    assert [c.name for c in table.tableColumns] == ["Date", "Diff", "cli"]
    assert [c.id for c in table.tableColumns] == [1, 2, 3]


def test_add_cli_column_does_not_backfill_existing_rows(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.add_cli_column(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert [ws2.cell(r, 3).value for r in (2, 3, 4)] == [None, None, None]


def test_the_saved_workbook_still_parses(pre_migration_book):
    """The nearest a test can get to Excel's own validation."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.move_comments(ws)
    migrate.add_cli_column(ws)
    sheet.save(wb, pre_migration_book)

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert sheet.has_cli_column(ws2) is True
    assert ws2["E1"].value == "Dage siden sidste skift"
    assert len(sheet.read_dates(pre_migration_book, SHEET)) == 3


def test_plan_lists_both_jobs_before_migration(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    steps = migrate.plan(ws)
    assert len(steps) == 2
    assert any("D -> E" in s for s in steps)
    assert any("cli" in s for s in steps)


def test_plan_is_empty_after_migration(post_migration_book):
    wb, ws = sheet.open_for_write(post_migration_book, SHEET)
    assert migrate.plan(ws) == []


def test_backup_copies_the_file_beside_the_original(pre_migration_book):
    made = migrate.backup(pre_migration_book)
    assert made.exists()
    assert made.parent == pre_migration_book.parent
    assert made.stat().st_size == pre_migration_book.stat().st_size
    assert made.name != pre_migration_book.name


def test_main_is_a_no_op_when_already_migrated(post_migration_book, monkeypatch, capsys):
    before_bytes = post_migration_book.read_bytes()
    calls = _spy_on_save(monkeypatch)

    monkeypatch.setenv("EXCEL_PATH", str(post_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    assert migrate.main([]) == 0
    assert "allerede" in capsys.readouterr().out.lower()
    # A true no-op: sheet.save is never called, and nothing on disk moved.
    assert calls == []
    assert post_migration_book.read_bytes() == before_bytes


def test_main_aborts_when_column_c_holds_something_else(pre_migration_book, monkeypatch, capsys):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    ws["C1"] = "noget andet"
    sheet.save(wb, pre_migration_book)
    before_bytes = pre_migration_book.read_bytes()
    calls = _spy_on_save(monkeypatch)

    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    assert migrate.main([]) == 1
    assert "C1" in capsys.readouterr().err
    # This guard fires BEFORE backup() runs -- a regression that saved here
    # would overwrite the original with no backup in existence. sheet.save
    # must never be called on this path (checked directly, not only via
    # bytes: an unmodified re-save can be byte-identical to the original
    # within the same wall-clock second, which would let this slip through
    # a bytes-only check).
    assert calls == []
    assert pre_migration_book.read_bytes() == before_bytes


def test_main_aborts_when_column_e_is_occupied(pre_migration_book, monkeypatch, capsys):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    ws["E9"] = "i vejen"
    sheet.save(wb, pre_migration_book)
    before_bytes = pre_migration_book.read_bytes()
    calls = _spy_on_save(monkeypatch)

    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    assert migrate.main([]) == 1
    assert "E" in capsys.readouterr().err
    # Same guard, same property, same reason for the call-spy over a
    # bytes-only check: it runs before backup(), so a regression that saved
    # here would destroy the original with nothing to restore it from.
    assert calls == []
    assert pre_migration_book.read_bytes() == before_bytes


def test_main_declined_changes_nothing(pre_migration_book, monkeypatch):
    before_bytes = pre_migration_book.read_bytes()
    calls = _spy_on_save(monkeypatch)

    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    monkeypatch.setattr("builtins.input", lambda _: "n")

    assert migrate.main([]) == 0
    ws = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert ws["C1"].value is None
    assert ws["D1"].value == "Dage siden sidste skift"
    # Declining must leave the workbook byte-identical -- no backup, no save.
    assert calls == []
    assert pre_migration_book.read_bytes() == before_bytes


def test_main_accepted_applies_both_jobs_and_backs_up(pre_migration_book, monkeypatch, capsys):
    before_bytes = pre_migration_book.read_bytes()
    calls = _spy_on_save(monkeypatch)

    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    monkeypatch.setattr("builtins.input", lambda _: "y")

    assert migrate.main([]) == 0

    # Exactly one save, against the real path -- proves the earlier abort
    # tests' calls == [] is pinning something real and not vacuously true.
    assert calls == [pre_migration_book]

    ws = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert ws["C1"].value == "cli"
    assert ws["E1"].value == "Dage siden sidste skift"
    assert ws["D1"].value is None
    backups = list(pre_migration_book.parent.glob("pre.*.bak.xlsx"))
    assert len(backups) == 1
    # The backup must hold the PRE-migration state -- taken before any
    # modification, not a copy of the now-migrated file.
    assert backups[0].read_bytes() == before_bytes


def test_main_yes_flag_skips_the_prompt(pre_migration_book, monkeypatch):
    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)

    def no_input(_):
        raise AssertionError("should not have prompted")

    monkeypatch.setattr("builtins.input", no_input)
    assert migrate.main(["-y"]) == 0
    assert openpyxl.load_workbook(pre_migration_book)[SHEET]["C1"].value == "cli"


def test_main_leaves_the_file_untouched_when_move_comments_raises_mid_loop(
        pre_migration_book, monkeypatch, capsys):
    """The critical safety property carried from Task 5's review: a raise
    from move_comments must abort BEFORE any save, so a half-migrated
    worksheet in memory never gets persisted over the file. move_comments is
    monkeypatched to raise directly -- constructing the raise through real
    column contents always trips the earlier `_blocked` guard instead (it
    refuses to even start once both D and E hold values), so this is the
    only way to exercise a raise that happens *after* the prompt/backup but
    *before* the save. main() must catch it, report on stderr (pointing at
    the backup that was already written), return 1 rather than letting the
    exception escape as a traceback, and -- above all -- never call
    sheet.save() afterwards."""
    def boom(_ws):
        raise RuntimeError("kolonne E har allerede en værdi i række 9")

    monkeypatch.setattr(migrate, "move_comments", boom)

    before_bytes = pre_migration_book.read_bytes()

    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    monkeypatch.setattr("builtins.input", lambda _: "y")

    assert migrate.main([]) == 1
    assert capsys.readouterr().err.strip() != ""
    assert pre_migration_book.read_bytes() == before_bytes


def test_main_only_adds_cli_column_when_comments_already_moved(
        pre_migration_book, monkeypatch, capsys):
    """Partial state: comments already live in E, cli still missing --
    move_comments must not run again. Critically, it must not disturb
    column widths on a D/E pair it has nothing to move: move_comments'
    trailing column-dimension block runs even when comment_rows(BEFORE) is
    empty, so calling it unconditionally would silently overwrite E's width
    with D's (or wipe a width the user set on D since the last migration)
    on a run whose printed plan only ever mentioned the cli column."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.move_comments(ws)
    sheet.save(wb, pre_migration_book)

    # Simulate a user who has since set a width on the now-empty D.
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    ws.column_dimensions["D"].width = 12.0
    sheet.save(wb, pre_migration_book)

    before = openpyxl.load_workbook(pre_migration_book)[SHEET]
    e_width_before = before.column_dimensions["E"].width
    assert e_width_before is not None

    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    monkeypatch.setattr("builtins.input", lambda _: "y")

    assert migrate.main([]) == 0
    out = capsys.readouterr().out
    assert "flyttet" not in out  # nothing was moved this run
    assert "cli" in out

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert ws2["C1"].value == "cli"
    assert ws2.column_dimensions["D"].width == pytest.approx(12.0)
    assert ws2.column_dimensions["E"].width == pytest.approx(e_width_before)


def test_main_only_moves_comments_when_cli_already_present(
        pre_migration_book, monkeypatch, capsys):
    """Partial state: cli already added, comments still in D -- only the
    move is outstanding, and the success message must say so."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    migrate.add_cli_column(ws)
    sheet.save(wb, pre_migration_book)

    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    monkeypatch.setattr("builtins.input", lambda _: "y")

    assert migrate.main([]) == 0
    out = capsys.readouterr().out
    assert "flyttet" in out
    assert "tilføjet" not in out  # the cli column was not touched this run

    ws2 = openpyxl.load_workbook(pre_migration_book)[SHEET]
    assert ws2["C1"].value == "cli"
    assert ws2["E1"].value == "Dage siden sidste skift"
    assert ws2["D1"].value is None


def test_add_cli_column_derives_both_edges_from_the_ref(pre_migration_book):
    """table.ref is the source of truth for both edges -- not a hardcoded
    A1 origin, and not fragile against absolute ($) anchors, which Excel
    itself writes on manual edits and which naive digit-stripping would
    mangle if the origin were not A1."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    table = ws.tables["Table2"]
    table.ref = f"$A$2:$B${TABLE_BOTTOM}"
    table.autoFilter.ref = table.ref

    migrate.add_cli_column(ws)
    sheet.save(wb, pre_migration_book)

    table2 = openpyxl.load_workbook(pre_migration_book)[SHEET].tables["Table2"]
    assert table2.ref == f"A2:C{TABLE_BOTTOM}"
    assert table2.autoFilter.ref == f"A2:C{TABLE_BOTTOM}"


def test_e_guard_does_not_fire_on_a_style_only_cell(pre_migration_book):
    """Mirrors the real workbook exactly: E2 exists in the sheet XML with a
    style but no value. Both comment_rows (which move_comments and _blocked
    both rely on) and _blocked itself must treat it as unoccupied -- this is
    precisely the scenario that would silently block the live migration if
    the guard were ever changed to test cell existence rather than value."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    assert (2, 5) in ws._cells  # the style-only cell really is there
    assert migrate.comment_rows(ws, migrate.COMMENT_COL_AFTER) == []
    assert migrate._blocked(ws) is None


def test_main_reports_a_missing_file(tmp_path, monkeypatch, capsys):
    missing = tmp_path / "nope.xlsx"
    monkeypatch.setenv("EXCEL_PATH", str(missing))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)

    assert migrate.main([]) == 1
    assert "findes ikke" in capsys.readouterr().err


def test_main_reports_a_missing_sheet(pre_migration_book, monkeypatch, capsys):
    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", "Findes Ikke")

    assert migrate.main([]) == 1
    err = capsys.readouterr().err
    assert "Findes Ikke" in err
    assert str(pre_migration_book) in err


def test_main_reports_a_locked_workbook(pre_migration_book, monkeypatch, capsys):
    def locked(wb, path):
        raise sheet.WorkbookLocked(path)

    monkeypatch.setattr(sheet, "save", locked)
    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    monkeypatch.setattr("builtins.input", lambda _: "y")

    assert migrate.main([]) == 1
    assert "Luk Excel" in capsys.readouterr().err


def test_main_reports_a_corrupt_workbook_and_points_at_the_backup_convention(
        tmp_path, monkeypatch, capsys):
    """The only way this tool can leave a corrupt workbook is an interrupted
    wb.save() -- and the user's natural next move is to re-run migrate. A
    truncated/garbage .xlsx is enough to trigger zipfile.BadZipFile, which
    must be reported in Danish rather than escape as a traceback."""
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"not a zip file at all, just garbage bytes")
    monkeypatch.setenv("EXCEL_PATH", str(corrupt))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)

    assert migrate.main([]) == 1
    err = capsys.readouterr().err
    assert "kan ikke læses" in err
    assert ".bak" in err


def test_main_reports_a_directory_path_as_an_unreadable_workbook(
        tmp_path, monkeypatch, capsys):
    """A directory raises openpyxl.utils.exceptions.InvalidFileException,
    not BadZipFile -- both must be handled."""
    monkeypatch.setenv("EXCEL_PATH", str(tmp_path))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)

    assert migrate.main([]) == 1
    assert "kan ikke læses" in capsys.readouterr().err


def test_blocked_does_not_create_c1_when_it_is_absent(pre_migration_book):
    """_blocked must be able to ask "does C1 hold something unexpected"
    without bringing C1 into existence as a side effect -- ws.cell() does
    that, which is exactly the hazard comment_rows' docstring warns about.
    On the real file C1 does not exist before migration."""
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    assert (1, sheet.CLI_COL) not in ws._cells

    assert migrate._blocked(ws) is None
    assert (1, sheet.CLI_COL) not in ws._cells

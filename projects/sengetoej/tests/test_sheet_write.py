import datetime as dt

import openpyxl
import pytest
from openpyxl.worksheet.formula import ArrayFormula

from sengetoej import sheet

from .conftest import COUNTER_TEXT, SHEET


def test_has_cli_column_is_false_before_migration(pre_migration_book):
    wb, ws = sheet.open_for_write(pre_migration_book, SHEET)
    assert sheet.has_cli_column(ws) is False


def test_has_cli_column_is_true_after_migration(post_migration_book):
    wb, ws = sheet.open_for_write(post_migration_book, SHEET)
    assert sheet.has_cli_column(ws) is True


def test_has_cli_column_ignores_case_and_whitespace(post_migration_book):
    wb, ws = sheet.open_for_write(post_migration_book, SHEET)
    ws.cell(1, sheet.CLI_COL, "  CLI ")
    assert sheet.has_cli_column(ws) is True


def test_dates_and_next_row_finds_the_first_empty_row(post_migration_book):
    wb, ws = sheet.open_for_write(post_migration_book, SHEET)
    dates, next_row = sheet.dates_and_next_row(ws)
    assert dates[-1] == dt.date(2026, 2, 3)
    assert next_row == 5  # rows 2-4 hold the three dates


def test_append_writes_the_date_and_the_cli_flag(post_migration_book):
    wb, ws = sheet.open_for_write(post_migration_book, SHEET)
    sheet.append_entry(ws, dt.date(2026, 3, 1), 5)
    sheet.save(wb, post_migration_book)

    ws2 = openpyxl.load_workbook(post_migration_book)[SHEET]
    assert ws2["A5"].value == dt.datetime(2026, 3, 1)
    assert ws2["C5"].value == 1


def test_append_copies_the_number_format_from_the_row_above(post_migration_book):
    wb, ws = sheet.open_for_write(post_migration_book, SHEET)
    sheet.append_entry(ws, dt.date(2026, 3, 1), 5)
    sheet.save(wb, post_migration_book)

    ws2 = openpyxl.load_workbook(post_migration_book)[SHEET]
    assert ws2["A5"].number_format == ws2["A4"].number_format == "mm-dd-yy"


def test_append_leaves_the_diff_formula_alone(post_migration_book):
    wb, ws = sheet.open_for_write(post_migration_book, SHEET)
    before = ws["B5"].value
    sheet.append_entry(ws, dt.date(2026, 3, 1), 5)
    sheet.save(wb, post_migration_book)

    ws2 = openpyxl.load_workbook(post_migration_book)[SHEET]
    assert ws2["B5"].value == before
    assert "Table2[[#This Row],[Date]]" in ws2["B5"].value


def test_append_leaves_the_comment_column_untouched(post_migration_book):
    wb, ws = sheet.open_for_write(post_migration_book, SHEET)
    sheet.append_entry(ws, dt.date(2026, 3, 1), 5)
    sheet.save(wb, post_migration_book)

    ws2 = openpyxl.load_workbook(post_migration_book)[SHEET]
    assert ws2["E1"].value == "Dage siden sidste skift"
    assert isinstance(ws2["E2"].value, ArrayFormula)
    assert ws2["E2"].value.text == COUNTER_TEXT


def test_append_does_not_change_the_table_ref(post_migration_book):
    wb, ws = sheet.open_for_write(post_migration_book, SHEET)
    before = ws.tables["Table2"].ref
    sheet.append_entry(ws, dt.date(2026, 3, 1), 5)
    sheet.save(wb, post_migration_book)

    ws2 = openpyxl.load_workbook(post_migration_book)[SHEET]
    assert ws2.tables["Table2"].ref == before


def test_save_reports_a_locked_workbook(post_migration_book, monkeypatch):
    wb, ws = sheet.open_for_write(post_migration_book, SHEET)

    def boom(*args, **kwargs):
        raise PermissionError(13, "Permission denied")

    monkeypatch.setattr(wb, "save", boom)
    with pytest.raises(sheet.WorkbookLocked):
        sheet.save(wb, post_migration_book)


def test_open_for_write_reports_a_missing_sheet(post_migration_book):
    with pytest.raises(sheet.SheetMissing):
        sheet.open_for_write(post_migration_book, "Findes Ikke")

import datetime as dt

import openpyxl
import pytest

from sengetoej import cli

from .conftest import SHEET


@pytest.fixture
def book_env(post_migration_book, monkeypatch):
    monkeypatch.setenv("EXCEL_PATH", str(post_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    monkeypatch.setattr(cli, "today", lambda: dt.date(2026, 2, 10))
    return post_migration_book


def rows(path):
    ws = openpyxl.load_workbook(path)[SHEET]
    return [(ws.cell(r, 1).value, ws.cell(r, 3).value) for r in range(2, 7)]


def test_new_with_a_date_appends_after_confirmation(book_env, monkeypatch, capsys):
    monkeypatch.setattr("builtins.input", lambda _: "y")
    assert cli.main(["--new", "09/02/2026"]) == 0

    assert rows(book_env)[3] == (dt.datetime(2026, 2, 9), 1)
    assert "Tilføjet i række 5." in capsys.readouterr().out


def test_new_without_a_date_uses_today(book_env, monkeypatch):
    monkeypatch.setattr("builtins.input", lambda _: "y")
    assert cli.main(["--new"]) == 0
    assert rows(book_env)[3][0] == dt.datetime(2026, 2, 10)


def test_the_prompt_shows_the_date_and_the_interval(book_env, monkeypatch, capsys):
    seen = {}

    def fake_input(prompt):
        seen["prompt"] = prompt
        return "y"

    monkeypatch.setattr("builtins.input", fake_input)
    cli.main(["--new", "09/02/2026"])
    assert "09/02/2026" in seen["prompt"]
    assert "6 dage siden sidste skift" in seen["prompt"]


def test_declining_writes_nothing_and_exits_zero(book_env, monkeypatch, capsys):
    monkeypatch.setattr("builtins.input", lambda _: "n")
    assert cli.main(["--new", "09/02/2026"]) == 0
    assert rows(book_env)[3] == (None, None)
    assert "Afbrudt." in capsys.readouterr().out


def test_yes_flag_skips_the_prompt(book_env, monkeypatch):
    def no_input(_):
        raise AssertionError("should not have prompted")

    monkeypatch.setattr("builtins.input", no_input)
    assert cli.main(["--new", "09/02/2026", "-y"]) == 0
    assert rows(book_env)[3][0] == dt.datetime(2026, 2, 9)


def test_a_malformed_date_is_a_usage_error(book_env, capsys):
    assert cli.main(["--new", "2026-02-09", "-y"]) == 2
    out, err = capsys.readouterr()
    assert "dd/mm/yyyy" in err
    assert "Åbner regnearket" not in out
    assert rows(book_env)[3] == (None, None)


def test_leading_zeros_are_optional(book_env, monkeypatch):
    monkeypatch.setattr("builtins.input", lambda _: "y")
    assert cli.main(["--new", "9/2/2026"]) == 0
    assert rows(book_env)[3][0] == dt.datetime(2026, 2, 9)


def test_a_future_date_is_rejected(book_env, capsys):
    assert cli.main(["--new", "01/03/2026", "-y"]) == 2
    out, err = capsys.readouterr()
    assert "fremtiden" in err
    assert "Åbner regnearket" not in out
    assert rows(book_env)[3] == (None, None)


def test_a_duplicate_date_is_rejected(book_env, capsys):
    assert cli.main(["--new", "03/02/2026", "-y"]) == 2
    err = capsys.readouterr().err
    assert "findes allerede" in err
    assert rows(book_env)[3] == (None, None)


def test_a_date_that_is_both_a_duplicate_and_earlier_reports_as_duplicate(
        book_env, capsys):
    # 15/01/2026 is DATES[1]: it duplicates an existing row AND is strictly
    # earlier than the last row (03/02/2026). The duplicate check must win.
    assert cli.main(["--new", "15/01/2026", "-y"]) == 2
    err = capsys.readouterr().err
    assert "findes allerede" in err
    assert "ligger før sidste række" not in err
    assert rows(book_env)[3] == (None, None)


def test_an_earlier_date_is_rejected(book_env, capsys):
    assert cli.main(["--new", "20/01/2026", "-y"]) == 2
    err = capsys.readouterr().err
    assert "ligger før sidste række" in err
    assert "03/02/2026" in err
    assert rows(book_env)[3] == (None, None)


def test_a_missing_cli_column_refuses_and_points_at_the_migration(
        pre_migration_book, monkeypatch, capsys):
    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    monkeypatch.setattr(cli, "today", lambda: dt.date(2026, 2, 10))

    assert cli.main(["--new", "09/02/2026", "-y"]) == 1
    err = capsys.readouterr().err
    assert "cli" in err
    assert "migrate" in err


def test_new_reports_a_missing_file(tmp_path, monkeypatch, capsys):
    monkeypatch.setenv("EXCEL_PATH", str(tmp_path / "nope.xlsx"))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    monkeypatch.setattr(cli, "today", lambda: dt.date(2026, 2, 10))

    assert cli.main(["--new", "-y"]) == 1
    assert "findes ikke" in capsys.readouterr().err


def test_new_reports_a_missing_sheet(pre_migration_book, monkeypatch, capsys):
    monkeypatch.setenv("EXCEL_PATH", str(pre_migration_book))
    monkeypatch.setenv("SENGETOEJ_SHEET", "Findes Ikke")
    monkeypatch.setattr(cli, "today", lambda: dt.date(2026, 2, 10))

    assert cli.main(["--new", "-y"]) == 1
    err = capsys.readouterr().err
    assert "Findes Ikke" in err
    assert str(pre_migration_book) in err


def test_new_reports_a_corrupt_workbook(tmp_path, monkeypatch, capsys):
    """The only way this tool can leave a corrupt workbook is an interrupted
    wb.save(), and the user's natural next move after --new fails mid-save
    is to just run --new again. A truncated/garbage .xlsx triggers
    zipfile.BadZipFile, which must be reported in Danish, not as a
    traceback."""
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"not a zip file at all, just garbage bytes")
    monkeypatch.setenv("EXCEL_PATH", str(corrupt))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    monkeypatch.setattr(cli, "today", lambda: dt.date(2026, 2, 10))

    assert cli.main(["--new", "-y"]) == 1
    assert "kan ikke læses" in capsys.readouterr().err


def test_a_locked_workbook_is_reported(book_env, monkeypatch, capsys):
    def locked(wb, path):
        raise cli.sheet.WorkbookLocked(path)

    monkeypatch.setattr(cli.sheet, "save", locked)
    assert cli.main(["--new", "09/02/2026", "-y"]) == 1
    assert "Luk Excel" in capsys.readouterr().err


def test_new_on_an_empty_sheet_appends_the_first_row(tmp_path, monkeypatch):
    from openpyxl.worksheet.table import Table
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"], ws["B1"], ws["C1"] = "Date", "Diff", "cli"
    ws.add_table(Table(displayName="Table2", ref="A1:C400"))
    wb.save(path)

    monkeypatch.setenv("EXCEL_PATH", str(path))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
    monkeypatch.setattr(cli, "today", lambda: dt.date(2026, 2, 10))

    assert cli.main(["--new", "-y"]) == 0
    ws2 = openpyxl.load_workbook(path)[SHEET]
    assert ws2["A2"].value == dt.datetime(2026, 2, 10)
    assert ws2["C2"].value == 1


def test_last_and_new_are_mutually_exclusive(capsys):
    with pytest.raises(SystemExit) as exc_info:
        cli.main(["--last", "3", "--new"])
    assert exc_info.value.code == 2
    assert "not allowed with argument" in capsys.readouterr().err

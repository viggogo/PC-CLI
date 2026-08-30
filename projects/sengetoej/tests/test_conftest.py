import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

from .conftest import COUNTER_TEXT, SHEET, TABLE_BOTTOM


def test_pre_migration_book_reproduces_the_traps(pre_migration_book):
    ws = openpyxl.load_workbook(pre_migration_book)[SHEET]
    table = ws.tables["Table2"]

    # The ref says nothing about where the data ends.
    assert table.ref == f"A1:B{TABLE_BOTTOM}"
    assert [c.name for c in table.tableColumns] == ["Date", "Diff"]

    # Diff formulas continue past the last date (row 4).
    assert ws["B40"].value is not None
    assert ws["A40"].value is None

    # C is empty; the comments are in D and outside the table.
    assert ws["C1"].value is None
    assert ws["D1"].value == "Dage siden sidste skift"

    # The counter is a styled array formula, not a plain string.
    counter = ws["D2"].value
    assert isinstance(counter, ArrayFormula)
    assert counter.text == COUNTER_TEXT
    assert counter.ref == "D2"
    assert ws["D2"].font.b is True


def test_post_migration_book_is_the_shape_migration_should_produce(post_migration_book):
    ws = openpyxl.load_workbook(post_migration_book)[SHEET]
    table = ws.tables["Table2"]

    assert table.ref == f"A1:C{TABLE_BOTTOM}"
    assert [c.name for c in table.tableColumns] == ["Date", "Diff", "cli"]
    assert ws["C1"].value == "cli"
    assert ws["D1"].value is None
    assert ws["E1"].value == "Dage siden sidste skift"
    assert isinstance(ws["E2"].value, ArrayFormula)
    assert ws["E2"].value.ref == "E2"

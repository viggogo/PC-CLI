"""Throwaway workbooks shaped like the real Sengetøj sheet.

The real sheet has three properties that break naive code, so the fixtures
reproduce all three:
  * Table2's ref spans the whole column (A1:B1048576), so it says nothing
    about where the data ends.
  * The Diff formula is pre-filled well past the last date, so a scan of
    column B would never terminate.
  * The comment column sits OUTSIDE the table and holds a styled
    ArrayFormula, which a value-only copy would silently mangle.
"""

import datetime as dt

import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

SHEET = "Sengetøj"

# Deliberately narrower than the real sheet's 1048576 so the tests stay fast.
# Everything that matters is that the ref extends far past the data.
TABLE_BOTTOM = 400
FORMULA_FILL_TO = 60

DATES = [dt.datetime(2026, 1, 1), dt.datetime(2026, 1, 15), dt.datetime(2026, 2, 3)]
COUNTER_TEXT = "=DAY(ABS(TODAY()-INDEX(A:A,COUNTA(A:A))))"


def _diff_formula(row: int) -> str:
    return (f'=IF(ISBLANK(Table2[[#This Row],[Date]]), " ", '
            f'Table2[[#This Row],[Date]]-A{row - 1})')


def _base_sheet(wb):
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "Date"
    ws["B1"] = "Diff"
    for i, d in enumerate(DATES, start=2):
        ws.cell(i, 1, d).number_format = "mm-dd-yy"
    for r in range(2, FORMULA_FILL_TO + 1):
        ws.cell(r, 2, _diff_formula(r))
    return ws


def _write_comments(ws, col: int):
    """Label, styled array formula, legend, and a per-row remark."""
    ws.cell(1, col, "Dage siden sidste skift")
    cell = ws.cell(2, col)
    cell.value = ArrayFormula(ref=f"{cell.column_letter}2", text=COUNTER_TEXT)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="FFFF00")
    ws.cell(3, col, "Obs: Ferie")
    ws.cell(4, col, "Dato indikerer at der er blevet skiftet sengetøj")
    ws.column_dimensions[ws.cell(1, col).column_letter].width = 20.6


def _add_table(ws, ref: str, names: list[str]):
    table = Table(displayName="Table2", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    # openpyxl's add_table does NOT derive tableColumns from the header row
    # (verified against the installed 3.1.5: Worksheet.add_table only stores
    # the Table object; nothing populates tableColumns). Without them the
    # file is invalid, so build them from the actual header cells within
    # `ref` -- which keeps the assert below a real consistency check between
    # the worksheet's header row and the names this fixture expects, not a
    # tautology. Setting tableColumns explicitly has a second consequence:
    # it suppresses openpyxl's save-time table initialisation, which is also
    # what synthesises the autoFilter -- so without setting one ourselves the
    # table would silently end up with autoFilter = None, losing a trap the
    # real Table2 has (its autoFilter spans the same over-wide ref as the
    # table itself).
    min_col, min_row, max_col, _ = range_boundaries(ref)
    table.tableColumns = [
        TableColumn(id=i, name=ws.cell(min_row, col).value)
        for i, col in enumerate(range(min_col, max_col + 1), start=1)
    ]
    table.autoFilter = AutoFilter(ref=ref)
    ws.add_table(table)
    assert [c.name for c in table.tableColumns] == names, (
        [c.name for c in table.tableColumns], names)
    return table


@pytest.fixture
def pre_migration_book(tmp_path):
    """Two-column Table2, comments in D, C empty. The state before migrate.py."""
    path = tmp_path / "pre.xlsx"
    wb = openpyxl.Workbook()
    ws = _base_sheet(wb)
    _write_comments(ws, 4)
    _add_table(ws, f"A1:B{TABLE_BOTTOM}", ["Date", "Diff"])
    wb.save(path)
    return path


@pytest.fixture
def post_migration_book(tmp_path):
    """Three-column Table2 with cli in C, comments in E, D empty."""
    path = tmp_path / "post.xlsx"
    wb = openpyxl.Workbook()
    ws = _base_sheet(wb)
    ws["C1"] = "cli"
    _write_comments(ws, 5)
    _add_table(ws, f"A1:C{TABLE_BOTTOM}", ["Date", "Diff", "cli"])
    wb.save(path)
    return path


@pytest.fixture
def header_only_book(tmp_path):
    """Three-column Table2 with cli in C, but zero data rows.

    The state a brand-new sheet is in before its first-ever entry: row 1 is
    the header, row 2 (FIRST_DATA_ROW) is blank. Exercises the FIRST_DATA_ROW
    edge case where "the row above" is the header, not a data row.
    """
    path = tmp_path / "header_only.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "Date"
    ws["B1"] = "Diff"
    ws["C1"] = "cli"
    _write_comments(ws, 5)
    _add_table(ws, f"A1:C{TABLE_BOTTOM}", ["Date", "Diff", "cli"])
    wb.save(path)
    return path


@pytest.fixture(autouse=True)
def never_touch_the_real_workbook(tmp_path, monkeypatch):
    """Point the tool's config at a throwaway path for EVERY test.

    install.ps1 writes a .env holding the REAL workbook path, and
    env.excel_path() falls back to that path even without one. A test that
    forgets to set EXCEL_PATH would therefore open the user's live
    spreadsheet. Individual tests still override these; this only makes the
    default safe.
    """
    monkeypatch.setenv("EXCEL_PATH", str(tmp_path / "unused.xlsx"))
    monkeypatch.setenv("SENGETOEJ_SHEET", SHEET)
